from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment


def make_excel(b2b, see, ids):
    wb = Workbook()
    create_b2b_sheet(b2b, wb)
    create_see_sheet(see, wb)
    create_result_sheet(ids, wb)
    create_ids_sheets(ids, wb)

    #    wb.save(filename="./diff-" + datetime.today().strftime('%Y%m%d%H%M%S') + ".xlsx")
    wb.save(filename="./diff.xlsx")  # for testing


def create_b2b_sheet(b2b, wb):
    b2b_sheet = wb.active
    b2b_sheet.title = "B2B"
    for val in b2b:
        b2b_sheet.append([val["bgm"], val["content"]])


def create_see_sheet(see, wb):
    see_sheet = wb.create_sheet(title="SEE")
    for val in see:
        see_sheet.append([val["bgm"], val["content"]])


def create_result_sheet(ids, wb):
    sheet = wb.create_sheet(title="Result")
    insert_header(sheet, ["Key", "Number of KO", "Result Counter"])
    insert_result_data(ids, sheet)
    set_result_style(ids, sheet)


def set_result_style(ids, sheet):
    sheet.column_dimensions['A'].width = 11
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    set_result_header_style(sheet)
    set_ok_ko_style(ids, sheet)
    set_borders_style(sheet, 'A1', 'C' + str(1 + len(ids)))


def set_ok_ko_style(ids, sheet):
    cell_range = sheet['C2':'C' + str(1 + len(ids))]
    for row in cell_range:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="C8F7C8" if cell.value == "OK" else "FFCCCB")
            cell.font = Font(color="0A5D00" if cell.value == "OK" else "800000")


def set_result_header_style(sheet):
    cell_range = sheet['A1':'C1']
    for row in cell_range:
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="00FFCC00")


def insert_result_data(ids, sheet):
    for val in ids:
        sheet.append([val["bgm"], val["ko_count"], val["unh_unt_result"]])


def set_borders_style(sheet, range1, range2):
    side = Side(border_style="medium")
    border = Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )
    cell_range = sheet[range1: range2]
    for row in cell_range:
        for cell in row:
            cell.border = border


def create_ids_sheets(ids, wb):
    for id in ids:
        create_id_sheet(id, wb)


def create_id_sheet(id, wb):
    sheet = wb.create_sheet(title="ID" + id["bgm"])
    make_header(sheet)
    make_diff(id, sheet)
    make_counters_check(id, sheet)
    make_ko_count(id, sheet)


def make_ko_count(id, sheet):
    insert_ko_count(id, sheet)
    set_ko_count_style(sheet)


def make_counters_check(id, sheet):
    insert_counters_data(id, sheet)
    set_counters_style(sheet)


def make_diff(id, sheet):
    insert_diff_data(id, sheet)
    set_diff_style(id, sheet)


def make_header(sheet):
    insert_header(sheet, ["B2B Message", "Seeburger Message", "Result", "Note",
                          "", "Segment", "Counter", "Result", "", "Numero KO"])
    set_ids_header_style(sheet)


def set_ko_count_style(sheet):
    set_borders_style(sheet, 'J1', 'J2')
    sheet.column_dimensions['J'].width = 13


def set_ids_header_style(sheet):
    cell_range = sheet['A1':'J1']
    for row in cell_range:
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="00FFCC00")
    sheet['E1'].fill = PatternFill(fill_type=None)
    sheet['I1'].fill = PatternFill(fill_type=None)


def set_counters_style(sheet):
    sheet.merge_cells('H2:H3')
    sheet['H2'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['H2'].fill = PatternFill("solid", fgColor="C8F7C8" if sheet['H2'].value == "OK" else "FFCCCB")
    sheet['H2'].font = Font(bold=True, color="0A5D00" if sheet['H2'].value == "OK" else "800000")
    set_borders_style(sheet, 'F1', 'H3')


def insert_ko_count(id, sheet):
    sheet['J2'] = id["ko_count"]


def insert_counters_data(id, sheet):
    sheet['F2'] = "UNH"
    sheet['F3'] = "UNT"
    sheet['G2'] = id["b2b_unh_counter"]
    sheet['G3'] = id["b2b_unt_counter"]
    sheet['H2'] = id["unh_unt_result"]


def set_diff_style(id, sheet):
    sheet.column_dimensions['A'].width = 45
    sheet.column_dimensions['B'].width = 45
    sheet.column_dimensions['C'].width = 7
    sheet.column_dimensions['D'].width = 25
    set_ok_ko_style(id["diff"], sheet)
    set_borders_style(sheet, 'A1', 'D' + str(1 + len(id["diff"])))


def insert_diff_data(id, sheet):
    for b2b_see_result in id["diff"]:
        sheet.append([b2b_see_result[0], b2b_see_result[1], b2b_see_result[2]])


def insert_header(sheet, titles):
    sheet.append(titles)
